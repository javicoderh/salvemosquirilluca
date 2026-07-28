"""
Análisis de duplicados e irregularidades en las firmas de la campaña.
Genera un informe Excel detallado con múltiples hojas.

Uso: python3 scripts/analizar_firmas.py
"""

import asyncio
import json
import os
import re
import sys
import urllib.request
import urllib.parse
import urllib.error
import base64
import time
import math
from datetime import datetime, timezone
from pathlib import Path
from collections import Counter, defaultdict

import asyncpg
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent

# ─── Utilidades ──────────────────────────────────────────────────────────────

def load_env():
    env = {}
    for path in (ROOT / ".env.local", ROOT / ".env"):
        if not path.exists():
            continue
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                eq = line.index("=") if "=" in line else -1
                if eq == -1:
                    continue
                key = line[:eq].strip()
                value = line[eq + 1:].strip()
                if value.startswith('"') and value.endswith('"'):
                    value = value[1:-1].replace("\\n", "\n")
                env[key] = value
    return env

ENV = load_env()

def ms_to_dt(ms):
    if not ms:
        return None
    try:
        return datetime.fromtimestamp(int(ms) / 1000, tz=timezone.utc)
    except Exception:
        return None

def ms_to_str(ms):
    dt = ms_to_dt(ms)
    return dt.strftime("%Y-%m-%d %H:%M:%S UTC") if dt else ""

# ─── PostgreSQL ───────────────────────────────────────────────────────────────

async def fetch_pg():
    url = ENV.get("POSTGRES_URL_NON_POOLING") or ENV.get("POSTGRES_PRISMA_URL") or ENV["POSTGRES_URL"]
    # asyncpg necesita la URL sin el channel_binding param
    url = re.sub(r"[?&]channel_binding=require", "", url)
    if "?" in url:
        url += "&ssl=require"
    else:
        url += "?ssl=require"

    print("Conectando a PostgreSQL...")
    conn = await asyncpg.connect(url)

    sigs = await conn.fetch("""
        SELECT
            id, first_name, last_name, full_name, rut, email, age,
            country, legal_nature, region, commune, affiliation, message,
            consent, updates, status, abuse_reason,
            risk_decision, risk_reasons, risk_score,
            source_ip_hash, source_user_agent_hash, source_fingerprint_hash,
            source_origin_host, source_submitted_at_ms,
            security_token_issued_at_ms, security_submit_time_ms,
            security_captcha_verified, security_high_protection_mode,
            dedupe_email_hash, dedupe_identity_hash, dedupe_rut_hash,
            dedupe_ip_email_hash,
            created_at_ms, updated_at_ms, source_name
        FROM signatures
        ORDER BY created_at_ms ASC
    """)

    foreign = await conn.fetch("""
        SELECT id, name, country, reason, email, email_hash,
               status, source_ip_hash, created_at_ms, updated_at_ms
        FROM foreign_signatures
        ORDER BY created_at_ms ASC
    """)

    await conn.close()
    print(f"  → {len(sigs)} firmas principales, {len(foreign)} extranjeras")
    return [dict(r) for r in sigs], [dict(r) for r in foreign]

# ─── Firestore REST ───────────────────────────────────────────────────────────

def get_firestore_token():
    import hashlib, hmac
    from cryptography.hazmat.primitives import hashes, serialization
    from cryptography.hazmat.primitives.asymmetric import padding
    from cryptography.hazmat.backends import default_backend

    header = base64.urlsafe_b64encode(json.dumps({"alg": "RS256", "typ": "JWT"}).encode()).rstrip(b"=").decode()
    now = int(time.time())
    payload_data = {
        "iss": ENV["FIREBASE_SERVICE_ACCOUNT_CLIENT_EMAIL"],
        "sub": ENV["FIREBASE_SERVICE_ACCOUNT_CLIENT_EMAIL"],
        "aud": "https://oauth2.googleapis.com/token",
        "iat": now,
        "exp": now + 3600,
        "scope": "https://www.googleapis.com/auth/datastore"
    }
    payload = base64.urlsafe_b64encode(json.dumps(payload_data).encode()).rstrip(b"=").decode()
    signing_input = f"{header}.{payload}".encode()

    private_key = serialization.load_pem_private_key(
        ENV["FIREBASE_SERVICE_ACCOUNT_PRIVATE_KEY"].encode(),
        password=None,
        backend=default_backend()
    )
    signature = private_key.sign(signing_input, padding.PKCS1v15(), hashes.SHA256())
    sig_b64 = base64.urlsafe_b64encode(signature).rstrip(b"=").decode()

    jwt = f"{header}.{payload}.{sig_b64}"

    data = urllib.parse.urlencode({
        "grant_type": "urn:ietf:params:oauth:grant-type:jwt-bearer",
        "assertion": jwt
    }).encode()
    req = urllib.request.Request("https://oauth2.googleapis.com/token", data=data)
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())["access_token"]

def parse_fs_value(v):
    if v is None:
        return None
    if "stringValue" in v:
        return v["stringValue"]
    if "integerValue" in v:
        return int(v["integerValue"])
    if "doubleValue" in v:
        return v["doubleValue"]
    if "booleanValue" in v:
        return v["booleanValue"]
    if "timestampValue" in v:
        return v["timestampValue"]
    if "mapValue" in v:
        fields = v["mapValue"].get("fields", {})
        return {k: parse_fs_value(val) for k, val in fields.items()}
    if "arrayValue" in v:
        return [parse_fs_value(i) for i in v["arrayValue"].get("values", [])]
    if "nullValue" in v:
        return None
    return str(v)

def fetch_firestore():
    project = ENV.get("FIREBASE_SERVICE_ACCOUNT_PROJECT_ID", ENV.get("PUBLIC_FIREBASE_PROJECT_ID"))
    collection = ENV.get("PUBLIC_FIREBASE_SIGNATURES_COLLECTION", "campaign_signatures")
    print("Obteniendo token de Firestore...")
    try:
        token = get_firestore_token()
    except Exception as e:
        print(f"  ✗ No se pudo autenticar con Firestore: {e}")
        return []

    base = f"https://firestore.googleapis.com/v1/projects/{project}/databases/(default)/documents/{collection}"
    docs = []
    page_token = None
    page = 1
    print(f"Descargando colección '{collection}'...")

    while True:
        params = {"pageSize": "300"}
        if page_token:
            params["pageToken"] = page_token
        url = base + "?" + urllib.parse.urlencode(params)
        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
        try:
            with urllib.request.urlopen(req) as resp:
                data = json.loads(resp.read())
        except Exception as e:
            print(f"  ✗ Error en página {page}: {e}")
            break

        batch = data.get("documents", [])
        docs.extend(batch)
        sys.stdout.write(f"\r  → Página {page}: {len(docs)} docs")
        sys.stdout.flush()
        page_token = data.get("nextPageToken")
        page += 1
        if not page_token:
            break

    print(f"\n  → Total Firestore: {len(docs)} docs")

    rows = []
    for doc in docs:
        fields = doc.get("fields", {})
        p = {k: parse_fs_value(v) for k, v in fields.items()}
        src = p.get("source", {}) or {}
        sec = p.get("security", {}) or {}
        rows.append({
            "id": doc["name"].split("/")[-1],
            "full_name": p.get("fullName", ""),
            "first_name": p.get("firstName", ""),
            "last_name": p.get("lastName", ""),
            "rut": p.get("rut", ""),
            "email": p.get("email", ""),
            "age": p.get("age", ""),
            "country": p.get("country", ""),
            "legal_nature": p.get("legalNature", ""),
            "region": p.get("region", ""),
            "commune": p.get("commune", ""),
            "affiliation": p.get("affiliation", ""),
            "message": p.get("message", ""),
            "status": p.get("status", ""),
            "risk_decision": p.get("riskDecision", ""),
            "risk_score": p.get("riskScore", 0),
            "risk_reasons": ", ".join(p.get("riskReasons", []) or []),
            "source_origin_host": src.get("originHost", ""),
            "source_ip_hash": src.get("ipHash", ""),
            "source_submitted_at_ms": src.get("submittedAtMs", 0),
            "security_submit_time_ms": sec.get("submitTimeMs", 0),
            "security_captcha_verified": sec.get("captchaVerified", False),
            "created_at_ms": p.get("createdAtMs", 0),
            "source_name": p.get("sourceName", ""),
        })
    return rows

# ─── Análisis ─────────────────────────────────────────────────────────────────

def normalize_rut(rut):
    """Quita puntos y espacios, lowercase."""
    if not rut:
        return ""
    return re.sub(r"[\s.\-]", "", str(rut)).lower().strip()

def normalize_email(email):
    if not email:
        return ""
    return str(email).lower().strip()

def normalize_name(name):
    if not name:
        return ""
    import unicodedata
    s = unicodedata.normalize("NFD", str(name).lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip()

DISPOSABLE_DOMAINS = {
    "mailinator.com", "tempmail.com", "guerrillamail.com", "10minutemail.com",
    "throwam.com", "yopmail.com", "sharklasers.com", "guerrillamailblock.com",
    "grr.la", "guerrillamail.info", "guerrillamail.biz", "guerrillamail.de",
    "guerrillamail.net", "guerrillamail.org", "spam4.me", "trashmail.com",
    "trashmail.me", "trashmail.net", "dispostable.com", "maildrop.cc",
    "mailnull.com", "spamgourmet.com", "spamgourmet.net", "getnada.com",
    "fakeinbox.com", "mailnesia.com", "discard.email", "spamherelots.com",
    "spamhereplease.com", "temporaryinbox.com", "mailexpire.com",
    "throwam.com", "tempr.email", "tempinbox.net", "email-temp.com",
    "emailondeck.com", "temp-mail.org", "tempmail.net", "fakemailgenerator.com",
}

def email_domain(email):
    if "@" in str(email):
        return str(email).split("@")[-1].lower().strip()
    return ""

VALID_REGIONS_CL = {
    "arica y parinacota", "tarapacá", "tarapaca", "antofagasta", "atacama",
    "coquimbo", "valparaíso", "valparaiso", "metropolitana", "o'higgins",
    "ohiggins", "maule", "ñuble", "nuble", "biobío", "biobio",
    "la araucanía", "la araucania", "los ríos", "los rios",
    "los lagos", "aysén", "aysen", "magallanes", "extranjero", "exterior",
    "no aplica", ""
}

SUSPICIOUS_NAME_PATTERNS = re.compile(
    r"^(test|prueba|admin|usuario|user|asdf|qwerty|aaaa|xxxx|nombre|apellido|"
    r"lorem|ipsum|ejemplo|sample|fake|falso|null|none|n/a|na|aaa|bbb|ccc|ddd|"
    r"xxx|zzz|abc|123|hola|hello)\b",
    re.IGNORECASE
)

def analyze_pg_signatures(sigs_raw):
    """Análisis completo de firmas de PostgreSQL."""
    df = pd.DataFrame(sigs_raw)
    if df.empty:
        return {}

    # Convertir timestamps
    df["fecha_envio"] = df["source_submitted_at_ms"].apply(ms_to_str)
    df["fecha_creacion"] = df["created_at_ms"].apply(ms_to_str)
    df["risk_reasons_str"] = df["risk_reasons"].apply(
        lambda x: ", ".join(x) if isinstance(x, list) else str(x or "")
    )

    results = {}

    # ── 1. Duplicados por email ──────────────────────────────────────────────
    df["email_norm"] = df["email"].apply(normalize_email)
    email_counts = df["email_norm"].value_counts()
    dup_emails = email_counts[email_counts > 1].index
    df_dup_email = df[df["email_norm"].isin(dup_emails)].copy()
    df_dup_email = df_dup_email.sort_values(["email_norm", "created_at_ms"])
    df_dup_email["veces_aparece"] = df_dup_email.groupby("email_norm")["email_norm"].transform("count")
    results["dup_email"] = df_dup_email[[
        "id", "full_name", "email", "rut", "country", "region", "status",
        "risk_score", "fecha_envio", "veces_aparece", "source_ip_hash"
    ]].rename(columns={
        "id": "ID", "full_name": "Nombre", "email": "Email", "rut": "RUT",
        "country": "País", "region": "Región", "status": "Estado",
        "risk_score": "Riesgo", "fecha_envio": "Fecha envío",
        "veces_aparece": "Nº duplicados", "source_ip_hash": "IP hash"
    })

    # ── 2. Duplicados por RUT ────────────────────────────────────────────────
    df["rut_norm"] = df["rut"].apply(normalize_rut)
    # Excluir RUTs vacíos o que sean solo "0"
    valid_ruts = df[df["rut_norm"].str.len() > 3]
    rut_counts = valid_ruts["rut_norm"].value_counts()
    dup_ruts = rut_counts[rut_counts > 1].index
    df_dup_rut = df[df["rut_norm"].isin(dup_ruts)].copy()
    df_dup_rut = df_dup_rut.sort_values(["rut_norm", "created_at_ms"])
    df_dup_rut["veces_aparece"] = df_dup_rut.groupby("rut_norm")["rut_norm"].transform("count")
    results["dup_rut"] = df_dup_rut[[
        "id", "full_name", "rut", "email", "country", "region", "status",
        "risk_score", "fecha_envio", "veces_aparece", "source_ip_hash"
    ]].rename(columns={
        "id": "ID", "full_name": "Nombre", "rut": "RUT", "email": "Email",
        "country": "País", "region": "Región", "status": "Estado",
        "risk_score": "Riesgo", "fecha_envio": "Fecha envío",
        "veces_aparece": "Nº duplicados", "source_ip_hash": "IP hash"
    })

    # ── 3. Duplicados por nombre completo normalizado ────────────────────────
    df["name_norm"] = df["full_name"].apply(normalize_name)
    name_counts = df[df["name_norm"].str.len() > 4]["name_norm"].value_counts()
    dup_names = name_counts[name_counts > 1].index
    df_dup_name = df[df["name_norm"].isin(dup_names)].copy()
    df_dup_name = df_dup_name.sort_values(["name_norm", "created_at_ms"])
    df_dup_name["veces_aparece"] = df_dup_name.groupby("name_norm")["name_norm"].transform("count")
    results["dup_nombre"] = df_dup_name[[
        "id", "full_name", "rut", "email", "country", "region", "status",
        "risk_score", "fecha_envio", "veces_aparece"
    ]].rename(columns={
        "id": "ID", "full_name": "Nombre", "rut": "RUT", "email": "Email",
        "country": "País", "region": "Región", "status": "Estado",
        "risk_score": "Riesgo", "fecha_envio": "Fecha envío",
        "veces_aparece": "Nº duplicados"
    })

    # ── 4. Mismo IP hash → múltiples firmas ─────────────────────────────────
    ip_counts = df["source_ip_hash"].value_counts()
    suspicious_ips = ip_counts[ip_counts >= 3].index
    df_ip = df[df["source_ip_hash"].isin(suspicious_ips)].copy()
    df_ip["firmas_desde_ip"] = df_ip.groupby("source_ip_hash")["source_ip_hash"].transform("count")
    df_ip = df_ip.sort_values(["firmas_desde_ip", "source_ip_hash", "created_at_ms"], ascending=[False, True, True])
    results["misma_ip"] = df_ip[[
        "id", "full_name", "email", "rut", "country", "region", "status",
        "risk_score", "fecha_envio", "firmas_desde_ip", "source_ip_hash"
    ]].rename(columns={
        "id": "ID", "full_name": "Nombre", "email": "Email", "rut": "RUT",
        "country": "País", "region": "Región", "status": "Estado",
        "risk_score": "Riesgo", "fecha_envio": "Fecha envío",
        "firmas_desde_ip": "Firmas desde misma IP", "source_ip_hash": "IP hash"
    })

    # ── 5. Mismo fingerprint → múltiples firmas ──────────────────────────────
    fp_counts = df["source_fingerprint_hash"].value_counts()
    suspicious_fps = fp_counts[fp_counts >= 3].index
    df_fp = df[df["source_fingerprint_hash"].isin(suspicious_fps)].copy()
    df_fp["firmas_desde_fp"] = df_fp.groupby("source_fingerprint_hash")["source_fingerprint_hash"].transform("count")
    df_fp = df_fp.sort_values(["firmas_desde_fp", "source_fingerprint_hash", "created_at_ms"], ascending=[False, True, True])
    results["mismo_fingerprint"] = df_fp[[
        "id", "full_name", "email", "rut", "country", "region", "status",
        "risk_score", "fecha_envio", "firmas_desde_fp", "source_fingerprint_hash"
    ]].rename(columns={
        "id": "ID", "full_name": "Nombre", "email": "Email", "rut": "RUT",
        "country": "País", "region": "Región", "status": "Estado",
        "risk_score": "Riesgo", "fecha_envio": "Fecha envío",
        "firmas_desde_fp": "Firmas mismo dispositivo", "source_fingerprint_hash": "Fingerprint hash"
    })

    # ── 6. Firmas flagged / rejected ─────────────────────────────────────────
    df_flag = df[df["status"].isin(["flagged", "rejected"])].copy()
    df_flag = df_flag.sort_values("risk_score", ascending=False)
    results["flagged_rejected"] = df_flag[[
        "id", "full_name", "email", "rut", "country", "region", "status",
        "risk_score", "risk_reasons_str", "fecha_envio",
        "security_submit_time_ms", "security_captcha_verified", "source_ip_hash"
    ]].rename(columns={
        "id": "ID", "full_name": "Nombre", "email": "Email", "rut": "RUT",
        "country": "País", "region": "Región", "status": "Estado",
        "risk_score": "Puntaje riesgo", "risk_reasons_str": "Razones riesgo",
        "fecha_envio": "Fecha envío",
        "security_submit_time_ms": "Tiempo envío (ms)",
        "security_captcha_verified": "Captcha OK",
        "source_ip_hash": "IP hash"
    })

    # ── 7. Envío ultra-rápido (< 8 segundos, pero aceptadas) ────────────────
    df["submit_sec"] = df["security_submit_time_ms"].apply(
        lambda x: round(int(x or 0) / 1000, 1)
    )
    df_fast = df[
        (df["submit_sec"] < 8) &
        (df["submit_sec"] > 0) &
        (df["status"] == "accepted")
    ].copy()
    df_fast = df_fast.sort_values("submit_sec")
    results["envio_rapido"] = df_fast[[
        "id", "full_name", "email", "rut", "country", "region",
        "submit_sec", "risk_score", "risk_reasons_str",
        "security_captcha_verified", "fecha_envio"
    ]].rename(columns={
        "id": "ID", "full_name": "Nombre", "email": "Email", "rut": "RUT",
        "country": "País", "region": "Región",
        "submit_sec": "Segundos en enviar",
        "risk_score": "Puntaje riesgo", "risk_reasons_str": "Razones riesgo",
        "security_captcha_verified": "Captcha OK", "fecha_envio": "Fecha envío"
    })

    # ── 8. Emails sospechosos ────────────────────────────────────────────────
    df["email_domain"] = df["email"].apply(email_domain)
    suspicious_email_mask = (
        df["email_domain"].isin(DISPOSABLE_DOMAINS) |
        df["email"].apply(lambda e: bool(re.search(r"\d{6,}", str(e)))) |
        df["email"].apply(lambda e: len(str(e)) < 6 or "@" not in str(e)) |
        df["email"].apply(lambda e: bool(re.match(r"^[a-z]{1,2}\d+@", str(e).lower())))
    )
    df_sus_email = df[suspicious_email_mask].copy()
    df_sus_email["motivo"] = df_sus_email.apply(lambda r: (
        ("dominio_desechable " if r["email_domain"] in DISPOSABLE_DOMAINS else "") +
        ("muchos_números " if re.search(r"\d{6,}", str(r["email"])) else "") +
        ("email_muy_corto " if len(str(r["email"])) < 6 or "@" not in str(r["email"]) else "") +
        ("patron_bot " if re.match(r"^[a-z]{1,2}\d+@", str(r["email"]).lower()) else "")
    ).strip(), axis=1)
    results["emails_sospechosos"] = df_sus_email[[
        "id", "full_name", "email", "email_domain", "rut", "country",
        "region", "status", "risk_score", "fecha_envio", "motivo"
    ]].rename(columns={
        "id": "ID", "full_name": "Nombre", "email": "Email",
        "email_domain": "Dominio", "rut": "RUT", "country": "País",
        "region": "Región", "status": "Estado", "risk_score": "Puntaje riesgo",
        "fecha_envio": "Fecha envío", "motivo": "Motivo sospecha"
    })

    # ── 9. RUTs sospechosos ──────────────────────────────────────────────────
    def rut_suspicious(rut_raw):
        r = normalize_rut(rut_raw)
        if not r:
            return "rut_vacío"
        if re.match(r"^0+k?$", r):
            return "rut_todo_ceros"
        if re.match(r"^1{5,}k?$", r):
            return "rut_repetido"
        if re.match(r"^(12345|11111|22222|33333|44444|55555|66666|77777|88888|99999)", r):
            return "rut_patron_simple"
        if re.match(r"^(test|prueba|fake|null|none|na)\d*", r, re.IGNORECASE):
            return "rut_texto_falso"
        # RUT extremadamente corto (< 6 dígitos antes del dígito verificador)
        digits = re.sub(r"[^0-9]", "", r)
        if len(digits) < 6:
            return "rut_demasiado_corto"
        # RUT con más de 9 dígitos (inválido en Chile)
        if len(digits) > 9:
            return "rut_demasiado_largo"
        return ""

    df["rut_sospecha"] = df["rut"].apply(rut_suspicious)
    df_sus_rut = df[df["rut_sospecha"] != ""].copy()
    results["ruts_sospechosos"] = df_sus_rut[[
        "id", "full_name", "rut", "email", "country", "region",
        "status", "risk_score", "fecha_envio", "rut_sospecha"
    ]].rename(columns={
        "id": "ID", "full_name": "Nombre", "rut": "RUT", "email": "Email",
        "country": "País", "region": "Región", "status": "Estado",
        "risk_score": "Puntaje riesgo", "fecha_envio": "Fecha envío",
        "rut_sospecha": "Motivo sospecha"
    })

    # ── 10. Mensajes duplicados (mismo mensaje, muchas firmas) ───────────────
    df["msg_norm"] = df["message"].apply(
        lambda x: re.sub(r"\s+", " ", str(x or "").lower().strip())
    )
    # Solo analizar mensajes no vacíos de más de 10 chars
    msg_df = df[df["msg_norm"].str.len() > 10]
    msg_counts = msg_df["msg_norm"].value_counts()
    dup_msgs = msg_counts[msg_counts >= 3].index
    df_dup_msg = df[df["msg_norm"].isin(dup_msgs)].copy()
    df_dup_msg["veces_mensaje"] = df_dup_msg.groupby("msg_norm")["msg_norm"].transform("count")
    df_dup_msg = df_dup_msg.sort_values(["veces_mensaje", "msg_norm"], ascending=[False, True])
    results["mensajes_duplicados"] = df_dup_msg[[
        "id", "full_name", "email", "country", "region", "status",
        "risk_score", "fecha_envio", "message", "veces_mensaje"
    ]].rename(columns={
        "id": "ID", "full_name": "Nombre", "email": "Email",
        "country": "País", "region": "Región", "status": "Estado",
        "risk_score": "Puntaje riesgo", "fecha_envio": "Fecha envío",
        "message": "Mensaje", "veces_mensaje": "Veces repetido"
    })

    # ── 11. Ráfagas: muchas firmas en ventana de 5 minutos ───────────────────
    df_sorted = df.sort_values("source_submitted_at_ms")
    df_sorted["ts"] = df_sorted["source_submitted_at_ms"].apply(
        lambda x: int(x or 0)
    )
    # Ventana de 5 min = 300000 ms
    window_ms = 5 * 60 * 1000
    bursts = []
    times = df_sorted["ts"].tolist()
    for i, t in enumerate(times):
        if t == 0:
            continue
        window_end = t + window_ms
        count_in_window = sum(1 for tt in times[i:] if t <= tt < window_end)
        if count_in_window >= 10:
            bursts.append({
                "ventana_inicio": ms_to_str(t),
                "ventana_fin": ms_to_str(window_end),
                "firmas_en_5min": count_in_window
            })

    # Desduplicar ventanas solapadas
    burst_df_raw = []
    if bursts:
        last_end = None
        for b in bursts:
            if last_end is None or b["ventana_inicio"] != last_end:
                burst_df_raw.append(b)
                last_end = b["ventana_inicio"]
        # Tomar las más grandes
        burst_df_raw = sorted(burst_df_raw, key=lambda x: -x["firmas_en_5min"])[:50]
    results["rafagas"] = pd.DataFrame(burst_df_raw) if burst_df_raw else pd.DataFrame(
        columns=["ventana_inicio", "ventana_fin", "firmas_en_5min"]
    )

    # ── 12. Nombres sospechosos ──────────────────────────────────────────────
    def name_suspicious(name):
        n = str(name or "").strip()
        if len(n) < 4:
            return "nombre_demasiado_corto"
        if SUSPICIOUS_NAME_PATTERNS.match(n):
            return "nombre_generico_falso"
        if re.match(r"^(.)\1{3,}", n.lower()):
            return "nombre_caracteres_repetidos"
        if re.search(r"\d{3,}", n):
            return "nombre_con_numeros"
        return ""

    df["nombre_sospecha"] = df["full_name"].apply(name_suspicious)
    df_sus_name = df[df["nombre_sospecha"] != ""].copy()
    results["nombres_sospechosos"] = df_sus_name[[
        "id", "full_name", "email", "rut", "country", "region",
        "status", "risk_score", "fecha_envio", "nombre_sospecha"
    ]].rename(columns={
        "id": "ID", "full_name": "Nombre", "email": "Email", "rut": "RUT",
        "country": "País", "region": "Región", "status": "Estado",
        "risk_score": "Puntaje riesgo", "fecha_envio": "Fecha envío",
        "nombre_sospecha": "Motivo sospecha"
    })

    return results, df


def analyze_cross_db(pg_df, fs_rows):
    """Detecta duplicados entre PostgreSQL y Firestore, con flag si el nombre difiere."""
    if not fs_rows:
        return pd.DataFrame(), 0, 0, 0

    fs_df = pd.DataFrame(fs_rows)
    fs_df["email_norm"] = fs_df["email"].apply(normalize_email)
    pg_df = pg_df.copy()
    pg_df["email_norm"] = pg_df["email"].apply(normalize_email)
    fs_df["rut_norm"] = fs_df["rut"].apply(normalize_rut)
    pg_df["rut_norm"] = pg_df["rut"].apply(normalize_rut)

    # Email en ambas
    fs_emails = set(fs_df["email_norm"].dropna())
    pg_emails = set(pg_df["email_norm"].dropna())
    shared_emails = fs_emails & pg_emails

    # RUT en ambas (solo RUTs con al menos 4 chars)
    valid_ruts_fs = set(fs_df[fs_df["rut_norm"].str.len() > 3]["rut_norm"])
    valid_ruts_pg = set(pg_df[pg_df["rut_norm"].str.len() > 3]["rut_norm"])
    shared_ruts = valid_ruts_fs & valid_ruts_pg

    cross_rows = []

    for email in shared_emails:
        pg_match = pg_df[pg_df["email_norm"] == email][["full_name","email","rut","status","created_at_ms","region","country"]].head(1)
        fs_match = fs_df[fs_df["email_norm"] == email][["full_name","email","rut","status","created_at_ms","region","country"]].head(1)
        if pg_match.empty or fs_match.empty:
            continue
        pg_r = pg_match.iloc[0]
        fs_r = fs_match.iloc[0]
        nombre_diferente = normalize_name(str(pg_r["full_name"])) != normalize_name(str(fs_r["full_name"]))
        cross_rows.append({
            "Tipo duplicado": "Email en ambas BDs",
            "⚠ Nombre distinto": "SÍ — REVISAR" if nombre_diferente else "",
            "Email / RUT compartido": email,
            "Nombre (PostgreSQL)": pg_r["full_name"],
            "RUT (PG)": pg_r["rut"],
            "País (PG)": pg_r["country"],
            "Región (PG)": pg_r["region"],
            "Estado (PG)": pg_r["status"],
            "Fecha firma (PG)": ms_to_str(pg_r["created_at_ms"]),
            "Nombre (Firestore)": fs_r["full_name"],
            "RUT (FS)": fs_r["rut"],
            "País (FS)": fs_r["country"],
            "Región (FS)": fs_r["region"],
            "Estado (Firestore)": fs_r["status"],
            "Fecha firma (Firestore)": ms_to_str(fs_r["created_at_ms"]),
        })

    for rut in shared_ruts:
        pg_match = pg_df[pg_df["rut_norm"] == rut][["full_name","email","rut","status","created_at_ms","region","country"]].head(1)
        fs_match = fs_df[fs_df["rut_norm"] == rut][["full_name","email","rut","status","created_at_ms","region","country"]].head(1)
        if pg_match.empty or fs_match.empty:
            continue
        pg_r = pg_match.iloc[0]
        fs_r = fs_match.iloc[0]
        nombre_diferente = normalize_name(str(pg_r["full_name"])) != normalize_name(str(fs_r["full_name"]))
        cross_rows.append({
            "Tipo duplicado": "RUT en ambas BDs",
            "⚠ Nombre distinto": "SÍ — REVISAR" if nombre_diferente else "",
            "Email / RUT compartido": rut,
            "Nombre (PostgreSQL)": pg_r["full_name"],
            "RUT (PG)": pg_r["rut"],
            "País (PG)": pg_r["country"],
            "Región (PG)": pg_r["region"],
            "Estado (PG)": pg_r["status"],
            "Fecha firma (PG)": ms_to_str(pg_r["created_at_ms"]),
            "Nombre (Firestore)": fs_r["full_name"],
            "RUT (FS)": fs_r["rut"],
            "País (FS)": fs_r["country"],
            "Región (FS)": fs_r["region"],
            "Estado (Firestore)": fs_r["status"],
            "Fecha firma (Firestore)": ms_to_str(fs_r["created_at_ms"]),
        })

    df_cross = pd.DataFrame(cross_rows)
    # Ordenar: primero los con nombre distinto (más sospechosos)
    if not df_cross.empty:
        df_cross = df_cross.sort_values(
            ["⚠ Nombre distinto", "Tipo duplicado"],
            ascending=[False, True]
        ).reset_index(drop=True)

    nombre_distinto_count = int((df_cross["⚠ Nombre distinto"] == "SÍ — REVISAR").sum()) if not df_cross.empty else 0
    return df_cross, len(shared_ruts), len(shared_emails), nombre_distinto_count


def build_summary(pg_results, pg_df, fs_rows, cross_ruts, cross_emails, cross_nombre_distinto):
    """Construye el resumen ejecutivo."""
    total_pg = len(pg_df)
    accepted = len(pg_df[pg_df["status"] == "accepted"])
    flagged = len(pg_df[pg_df["status"] == "flagged"])
    rejected = len(pg_df[pg_df["status"] == "rejected"])
    total_fs = len(fs_rows)
    total_ambas = total_pg + total_fs

    dup_email_count = len(pg_results["dup_email"]["Email"].unique()) if not pg_results["dup_email"].empty else 0
    dup_rut_count = len(pg_results["dup_rut"]["RUT"].unique()) if not pg_results["dup_rut"].empty else 0
    dup_name_count = len(pg_results["dup_nombre"]["Nombre"].unique()) if not pg_results["dup_nombre"].empty else 0
    suspicious_ip = len(pg_results["misma_ip"]) if not pg_results["misma_ip"].empty else 0
    flagged_total = len(pg_results["flagged_rejected"])
    fast_subs = len(pg_results["envio_rapido"])
    sus_email = len(pg_results["emails_sospechosos"])
    sus_rut = len(pg_results["ruts_sospechosos"])
    dup_msg = len(pg_results["mensajes_duplicados"])
    sus_name = len(pg_results["nombres_sospechosos"])

    rows = [
        ("TOTALES", "", ""),
        ("Total firmas combinadas (ambas BDs)", total_ambas, ""),
        ("  → PostgreSQL (sistema actual)", total_pg, ""),
        ("     → Aceptadas", accepted, f"{accepted/total_pg*100:.1f}%"),
        ("     → Marcadas (flagged)", flagged, f"{flagged/total_pg*100:.1f}%"),
        ("     → Rechazadas", rejected, f"{rejected/total_pg*100:.1f}%"),
        ("  → Firestore (históricas, antes del sistema actual)", total_fs, ""),
        ("", "", ""),
        ("DUPLICADOS INTERNOS — solo PostgreSQL", "", ""),
        ("Firmas con email duplicado dentro de PG", dup_email_count, "ver hoja 'Dup. por Email'"),
        ("Firmas con RUT duplicado dentro de PG", dup_rut_count, "ver hoja 'Dup. por RUT'"),
        ("Firmas con nombre duplicado dentro de PG", dup_name_count, "ver hoja 'Dup. por Nombre'"),
        ("", "", ""),
        ("DUPLICADOS ENTRE BDs (PostgreSQL ↔ Firestore)", "", ""),
        ("Personas con RUT repetido en ambas BDs", cross_ruts, "ver hoja 'Dup. entre BDs'"),
        ("Personas con email repetido en ambas BDs", cross_emails, "ver hoja 'Dup. entre BDs'"),
        ("  ⚠ De los anteriores, con NOMBRE DISTINTO (sospechoso)", cross_nombre_distinto, "ver hoja 'Dup. entre BDs' — columna '⚠ Nombre distinto'"),
        ("", "", ""),
        ("PATRONES SOSPECHOSOS — PostgreSQL", "", ""),
        ("Firmas desde misma IP (≥3 firmas/IP)", suspicious_ip, "ver hoja 'Misma IP'"),
        ("Firmas flagged o rechazadas", flagged_total, "ver hoja 'Flagged-Rejected'"),
        ("Firmas enviadas en < 8 segundos (aceptadas)", fast_subs, "ver hoja 'Envío Rápido'"),
        ("Emails con patrón sospechoso o desechable", sus_email, "ver hoja 'Emails Sospechosos'"),
        ("RUTs con formato inválido o sospechoso", sus_rut, "ver hoja 'RUTs Sospechosos'"),
        ("Mensajes repetidos (≥3 veces idéntico)", dup_msg, "ver hoja 'Mensajes Repetidos'"),
        ("Nombres con patrón genérico o falso", sus_name, "ver hoja 'Nombres Sospechosos'"),
        ("", "", ""),
        ("Generado el", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""),
    ]
    return pd.DataFrame(rows, columns=["Indicador", "Valor", "Detalle"])


# ─── Excel con formato ────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="DEEAF1")
WARNING_FILL = PatternFill("solid", fgColor="FFE699")
RED_FILL = PatternFill("solid", fgColor="FF9999")
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
SECTION_FONT = Font(color="FFFFFF", bold=True, size=11)

def style_sheet(ws, df, color_col=None, red_values=None, orange_values=None):
    """Aplica estilos básicos a una hoja."""
    # Cabecera
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Filas alternadas + colores condicionales
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical="center")
            if color_col is not None:
                col_letter = get_column_letter(cell.column)
                header_cell = ws[f"{col_letter}1"].value
                if header_cell == color_col:
                    val = str(cell.value or "")
                    if red_values and val in red_values:
                        cell.fill = RED_FILL
                        continue
                    if orange_values and val in orange_values:
                        cell.fill = WARNING_FILL
                        continue
            cell.fill = fill

    # Anchos automáticos
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value or ""))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

    ws.freeze_panes = "A2"


def write_excel(pg_results, cross_df, summary_df, pg_df, fs_rows, outpath):
    writer = pd.ExcelWriter(outpath, engine="openpyxl")

    # ── Resumen ──────────────────────────────────────────────────────────────
    summary_df.to_excel(writer, sheet_name="Resumen", index=False)

    sheets = [
        ("Dup. por Email",        pg_results["dup_email"],        "Estado", ["rejected"], ["flagged"]),
        ("Dup. por RUT",          pg_results["dup_rut"],          "Estado", ["rejected"], ["flagged"]),
        ("Dup. por Nombre",       pg_results["dup_nombre"],       "Estado", ["rejected"], ["flagged"]),
        ("Dup. entre BDs",        cross_df,  "⚠ Nombre distinto", ["SÍ — REVISAR"], None),
        ("Misma IP",              pg_results["misma_ip"],         "Estado", ["rejected"], ["flagged"]),
        ("Mismo Dispositivo",     pg_results["mismo_fingerprint"], "Estado", ["rejected"], ["flagged"]),
        ("Flagged-Rejected",       pg_results["flagged_rejected"], "Estado", ["rejected"], ["flagged"]),
        ("Envío Rápido",          pg_results["envio_rapido"],     None,     None,          None),
        ("Emails Sospechosos",    pg_results["emails_sospechosos"], "Estado", ["rejected"], ["flagged"]),
        ("RUTs Sospechosos",      pg_results["ruts_sospechosos"], "Estado", ["rejected"], ["flagged"]),
        ("Mensajes Repetidos",    pg_results["mensajes_duplicados"], "Estado", ["rejected"], ["flagged"]),
        ("Nombres Sospechosos",   pg_results["nombres_sospechosos"], "Estado", ["rejected"], ["flagged"]),
        ("Ráfagas de Envío",      pg_results["rafagas"],          None,     None,          None),
    ]

    for sheet_name, df_sheet, color_col, red_vals, orange_vals in sheets:
        if df_sheet is None or df_sheet.empty:
            df_sheet = pd.DataFrame({"Nota": ["Sin hallazgos en esta categoría"]})
        df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.close()

    # Aplicar estilos post-escritura
    wb = load_workbook(outpath)

    style_sheet(wb["Resumen"], summary_df)

    for sheet_name, df_sheet, color_col, red_vals, orange_vals in sheets:
        if sheet_name in wb.sheetnames:
            style_sheet(wb[sheet_name], df_sheet, color_col, red_vals, orange_vals)

    wb.save(outpath)
    print(f"✓ Estilos aplicados")


# ─── Main ─────────────────────────────────────────────────────────────────────

async def main():
    print("=== Análisis de duplicados e irregularidades ===\n")

    pg_sigs, pg_foreign = await fetch_pg()
    fs_rows = fetch_firestore()

    print("\nAnalizando datos...")
    pg_df = pd.DataFrame(pg_sigs)

    pg_results, pg_df_enriched = analyze_pg_signatures(pg_sigs)
    cross_df, cross_ruts, cross_emails, cross_nombre_distinto = analyze_cross_db(pg_df_enriched.copy(), fs_rows)
    summary_df = build_summary(pg_results, pg_df_enriched, fs_rows, cross_ruts, cross_emails, cross_nombre_distinto)

    # Imprimir resumen en terminal
    section_headers = {
        "TOTALES",
        "DUPLICADOS INTERNOS — solo PostgreSQL",
        "DUPLICADOS ENTRE BDs (PostgreSQL ↔ Firestore)",
        "PATRONES SOSPECHOSOS — PostgreSQL",
    }
    print("\n─── RESUMEN ─────────────────────────────────────────────────────")
    for _, row in summary_df.iterrows():
        if row["Indicador"] in section_headers:
            print(f"\n{row['Indicador']}")
        elif row["Indicador"] == "":
            pass
        else:
            val = f"{row['Valor']}"
            det = f"  [{row['Detalle']}]" if row["Detalle"] else ""
            print(f"  {row['Indicador']:<60} {val}{det}")
    print()

    today = datetime.now().strftime("%Y-%m-%d")
    outpath = ROOT / f"informe-firmas-{today}.xlsx"
    print(f"Generando informe Excel → {outpath}")
    write_excel(pg_results, cross_df, summary_df, pg_df_enriched, fs_rows, str(outpath))
    print(f"\n✓ Informe generado: {outpath}")


if __name__ == "__main__":
    asyncio.run(main())
